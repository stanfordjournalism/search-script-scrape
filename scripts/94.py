# The New York City high school with the highest average math score in the latest SAT results

# Notes:
# As this is one of the last exercises that I've written out, it includes code
#    that is both lazy and convoluted. For example, this is the first time
#    I've tried openpyxl as opposed to xlrd for reading Excel files
#    and it shows: http://openpyxl.readthedocs.org/en/latest/index.html
#
# You can check out other scraping/spreadsheet parsing examples
#  in the repo to find cleaner ways of doing this kind of task.
#
#
#
# Landing page:
# http://schools.nyc.gov/Accountability/data/TestResults/default.htm
#
## Relevant text from the webpage:
# The most recent school level results for New York City on the SAT.
#  Results are available at the school level for the graduating
#  seniors of 2014.  For a summary report of SAT, PSAT, and AP achievement
#  for 2014, please click here.
#
## Target URL looks like this:
# http://schools.nyc.gov/NR/rdonlyres/CE9139F0-9F3A-4C42-ACB8-74F2D014802F/
#                                           171380/2014SATWebsite10214.xlsx
#
# It's just as likely that by next year, they'll redesign or restructure
#  the site. So this scraping code is unstable. But it works as of August 2015.
import csv
import requests
from io import BytesIO
from operator import itemgetter
from urllib.parse import urljoin
from lxml import html
from openpyxl import load_workbook
LANDING_PAGE_URL = 'http://schools.nyc.gov/Accountability/data/TestResults/default.htm'

doc = html.fromstring(requests.get(LANDING_PAGE_URL).text)
# instead of using xpath, let's just use a sloppy csscelect
urls = [a.attrib.get('href') for a in doc.cssselect('a')]
# that awkward `get` is because not all anchor tags have hrefs...and
#  this is why we use xpath...
_xurl = next(url for url in urls if url and 'SAT' in url and 'xls' in url) # blargh
xlsx_url = urljoin(LANDING_PAGE_URL, _xurl)
print("Downloading", xlsx_url)
# download the spreadsheet...instead of writing to disk
# let's just keep it in memory and pass it directly to load_workbook()
xlsx = BytesIO(requests.get(xlsx_url).content)
wb = load_workbook(xlsx)
# The above command will print out a warning:
#     /site-packages/openpyxl/workbook/names/named_range.py:121: UserWarning:
#              Discarded range with reserved name
#              warnings.warn("Discarded range with reserved name")

### Dealing with the worksheet structure
# The 2014 edition contains two worksheets, the first being "Notes"
#  and the second being "2014 SAT Results"
# Let's write an agnostic function as if we didn't know how each year's
#   spreadsheet was actually structured
sheet = next(s for s in wb.worksheets if "results" in s.title.lower())
# I don't understand openpyxl's API so I'm just going to
#   practice nested list comprehensions
# Note that the first column is just an ID field which we don't care about
rows = [[cell.value for cell in row[1:]] for row in sheet.iter_rows()]
headers = rows[0]
# make it into a list of dicts
data = [dict(zip(headers, r)) for r in rows[1:]]
# I think we can assume that the header will change every year/file
# so let's write another agnostic iterating function to do a best guess
mathheader = next(h for h in headers if 'math' in h.lower())
# Not every school has a number for this column
mathschools = [d for d in data if isinstance(d[mathheader], int)]
topschool = max(mathschools, key = itemgetter(mathheader))
# since we've done so much work to get here, so
#   let's calculate the average of the averages -- which requires a
#    weighting of math score averages against number of SAT taker
#    and include that in the printed answer

# find the header that says '# of SAT Takers in 20XX':
numheader = next(h for h in headers if 'takers' in h.lower())
total_takers = sum(s[numheader] for s in mathschools)
mathsums = sum(s[mathheader] * s[numheader] for s in mathschools)
mathavg = mathsums // total_takers
tmp_answer = """{name} had the highest average SAT math score: {top_score}
    This was {diff_score} points higher than the city average of {avg_score}
"""
answer = tmp_answer.format(name = topschool['High School'],
    top_score = topschool[mathheader],
    diff_score = topschool[mathheader] - mathavg,
    avg_score = mathavg
)

print(answer)
# Output for 2014:
# STUYVESANT HIGH SCHOOL had the highest average SAT math score: 737
#    This was 272 points higher than the city average of 465
